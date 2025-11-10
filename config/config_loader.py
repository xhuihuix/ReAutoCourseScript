from openpyxl import load_workbook
from yaml import load, FullLoader
from dataclasses import dataclass
from typing import Optional


@dataclass
class WebConfig:
    base_domain: str
    sso_login_url: str
    client_id: str
    site_code: str
    qr_code_url: str
    check_login_status_url: str
    check_is_need_setting: str
    course_status_url: str
    select_elective_url: str
    project_class_id_url: str
    login_page_url: str
    redirect_url: str


@dataclass
class ProjectConfig:
    user_batch_size: int


@dataclass
class QrCodeConfig:
    api_url: str
    token: str


@dataclass
class VideoPlayConfig:
    class_id: str
    each_batch: int


@dataclass
class CookieConfig:
    save_path: str = "save_data/cookies"


@dataclass
class UserConfig:
    file_path: str = "config/账户信息.xlsx"


@dataclass
class AppConfig:
    web: WebConfig
    project: ProjectConfig
    qr_code: QrCodeConfig
    video_play: VideoPlayConfig
    cookie: CookieConfig
    account: UserConfig

@dataclass
class UserData:
    user_name: str
    need_credit: int
    username: str
    userpwd: str
    must_learn_course: list


class ConfigLoader:
    @staticmethod
    def load_config(config_path: str = "config/config.yaml") -> AppConfig:
        with open(config_path, "r", encoding="utf-8") as f:
            config_data = load(f, Loader=FullLoader)

        return AppConfig(
            web=WebConfig(**config_data["web"]),
            project=ProjectConfig(**config_data["project"]),
            qr_code=QrCodeConfig(**config_data["qr_code"]),
            video_play=VideoPlayConfig(**config_data["video_play"]),
            cookie=CookieConfig(**config_data["cookie"]),
            account=UserConfig(**config_data["account"])
        )


def read_user_info(user_config: UserConfig) -> list[UserData]:
    """
    新版本excel读取,数据都被存储在一个sheet里,依次为序号/姓名/账号/密码/指定学习项目
    :param user_config:
    :return:
    """
    # 检查配置是否有效
    if not hasattr(user_config, 'file_path') or not user_config.file_path:
        print("无用户路径信息, 请配置yaml文件")
        return []

    data_path = user_config.file_path

    # 检查文件是否存在
    import os
    if not os.path.exists(data_path):
        print(f"用户数据文件不存在: {data_path}")
        return []

    file_suffix = data_path.split(".")[-1].lower()

    if file_suffix in ["xlsx", "xls"]:
        try:
            # 使用 openpyxl 读取 xlsx 文件
            workbook = load_workbook(data_path)
        except Exception as e:
            print(f"无法打开Excel文件 {data_path}: {e}")
            return []

        if len(workbook.sheetnames) >= 1:
            user_data_sheet = workbook[workbook.sheetnames[0]]
        else:
            print("错误：未找到用户数据表")
            workbook.close()
            return []

        users_data = []
        for row_index, row in enumerate(user_data_sheet.iter_rows(min_row=2, values_only=True), 2):
            # 跳过空行
            if not any(cell is not None for cell in row):
                continue

            # 检查必要字段是否存在
            if len(row) < 5:
                print(f"警告：第{row_index}行数据不完整，跳过该行")
                continue

            # 处理用户名为空的情况
            username = row[2]
            if not username:
                print(f"警告：第{row_index}行用户名为空，跳过该行")
                continue

            try:
                user = UserData(
                    user_name=str(row[1]) if row[1] is not None else "",
                    username=str(username),
                    userpwd=str(row[3]) if row[3] is not None else "",
                    need_credit=0,
                    must_learn_course=str(row[4]).replace("，", ",").replace("\n", "").split(",")
                    if row[4] is not None else []
                )
                users_data.append(user)
            except Exception as e:
                print(f"警告：第{row_index}行数据处理出错: {e}，跳过该行")

        workbook.close()
        return users_data
    else:
        print(f"不支持的文件格式: {file_suffix}，请使用xlsx或xls格式")
        return []

def read_user_info_old(user_config: UserConfig) -> list[UserData]:
    if not hasattr(user_config, 'file_path'):
        print("无用户路径信息, 请配置yaml文件")
        return []

    data_path = user_config.file_path
    file_suffix = data_path.split(".")[-1]

    if file_suffix in ["xlsx"]:
        # 使用 openpyxl 读取 xlsx 文件
        workbook = load_workbook(data_path)

        # 明确第一个sheet为用户数据表
        if len(workbook.sheetnames) >= 1:
            user_data_sheet = workbook[workbook.sheetnames[0]]
        else:
            print("错误：未找到用户数据表")
            workbook.close()
            return []

        # 读取第三个sheet作为map_credit映射表
        if len(workbook.sheetnames) > 1:
            credit_sheet = workbook[workbook.sheetnames[2]]
            map_credit = {}
            # 从第二行开始读取映射关系（跳过标题行）
            for row in credit_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # 确保键值都存在
                    map_credit[str(row[0])] = int(row[1])
        else:
            print("警告：未找到第二个sheet，使用默认映射表")
            map_credit = {
                "15分": 15,
                "15分市级国家级": 15,
                "15分市级国家级6分": 15,
                "12分市级国家级各6分": 12,
                "10分国家级五分": 10,
                "5分市级5分国家级": 10,
            }
        # config_sheet = workbook[workbook.sheetnames[3]]
        # global_config = get_config()
        # config_row = list(config_sheet.iter_rows(min_row=2, max_row=3, values_only=True))
        # global_config.project.user_batch_size = int(config_row[0][1])
        # global_config.video_play.each_batch = int(config_row[1][1])
        #
        # exam_sheet = workbook[workbook.sheetnames[1]]

        users_data = []
        # 跳过标题行，从第二行开始读取用户数据
        for row in user_data_sheet.iter_rows(min_row=2, values_only=True):
            if str(row[6]) not in map_credit:
                print(f"ERROR 没有对应学分凭证数据 '{row[6]}'，请检查数据")
                continue
            user = UserData(
                city=str(row[1]),
                user_name=str(row[2]),
                major=str(row[3]),
                ic_card=str(row[4]),
                username=str(row[5]),
                need_credit=map_credit[str(row[6])],
                userpwd=str(row[8]),
                must_learn_course=str(row[9]).replace("，", ",").replace("\n","").split(",")
            )
            users_data.append(user)

        workbook.close()
        return users_data

    elif file_suffix in ["xls"]:
        # 如果需要支持 xls，可以添加 xlrd 支持
        print("暂不支持 .xls 格式，请使用 .xlsx 格式")
        return []

    return []


# 全局配置实例
config: Optional[AppConfig] = None


def get_config() -> AppConfig:
    global config
    if config is None:
        config = ConfigLoader.load_config()
    return config
